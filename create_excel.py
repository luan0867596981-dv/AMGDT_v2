import os
import glob
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_cell(cell, font_name="Segoe UI", font_size=10, bold=False, italic=False, color="000000", fill_color=None, alignment="left", border_style="thin", border_color="D3D3D3"):
    cell.font = Font(name=font_name, size=font_size, bold=bold, italic=italic, color=color)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    cell.alignment = Alignment(horizontal=alignment, vertical="center", wrap_text=True)
    
    thin = Side(border_style=border_style, color=border_color)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def main():
    root = os.path.dirname(os.path.abspath(__file__))
    tables_dir = os.path.join(root, 'results', 'tables')
    output_path = os.path.join(root, 'results', 'ket_qua_thuc_nghiem.xlsx')

    # Load all 6 files
    b_old = pd.read_csv(os.path.join(tables_dir, '10_fold_results_B-dataset-old.csv'))
    b_new = pd.read_csv(os.path.join(tables_dir, '10_fold_results_B-dataset.csv'))
    
    c_old = pd.read_csv(os.path.join(tables_dir, '10_fold_results_C-dataset-old.csv'))
    c_new = pd.read_csv(os.path.join(tables_dir, '10_fold_results_C-dataset.csv'))
    
    f_old = pd.read_csv(os.path.join(tables_dir, '10_fold_results_F-dataset-old.csv'))
    f_new = pd.read_csv(os.path.join(tables_dir, '10_fold_results_F-dataset.csv'))

    # Normalize Fold columns
    for df in [b_old, b_new, c_old, c_new, f_old, f_new]:
        # Identify first column name
        first_col = df.columns[0]
        df.rename(columns={first_col: 'Fold'}, inplace=True)
        # Normalize fold string names (e.g. 'Fold 1' or '1' -> 'Fold 1')
        df['Fold'] = df['Fold'].apply(lambda x: f"Fold {x}" if str(x).isdigit() else str(x))

    # Extract Mean values for Summary Table
    # Filter only Mean row
    def get_mean_row(df):
        row = df[df['Fold'].str.contains('Mean', case=False, na=False)]
        if len(row) > 0:
            return row.iloc[0]
        return None

    means = {
        'B': {'old': get_mean_row(b_old), 'new': get_mean_row(b_new)},
        'C': {'old': get_mean_row(c_old), 'new': get_mean_row(c_new)},
        'F': {'old': get_mean_row(f_old), 'new': get_mean_row(f_new)}
    }

    # Open workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ────────────────────────────────────────────────────────────────────────
    # SHEET 1: OVERVIEW COMPARISON
    # ────────────────────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Tổng quan So sánh")
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title
    ws_summary.merge_cells("A1:I1")
    ws_summary["A1"] = "BẢNG SO SÁNH HIỆU SUẤT MÔ HÌNH: BASELINE VS AMNTDDA"
    format_cell(ws_summary["A1"], font_size=14, bold=True, color="FFFFFF", fill_color="0D9488", alignment="center")
    ws_summary.row_dimensions[1].height = 40

    # Headers
    headers = ["Tập dữ liệu", "Mô hình", "AUC", "AUPR", "Accuracy", "Precision", "Recall", "F1-score", "MCC"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=2, column=col_idx, value=h)
        format_cell(cell, font_size=10, bold=True, color="FFFFFF", fill_color="115E59", alignment="center")
    ws_summary.row_dimensions[2].height = 25

    summary_rows = []
    for letter in ['B', 'C', 'F']:
        old_m = means[letter]['old']
        new_m = means[letter]['new']
        
        # Row for Baseline
        summary_rows.append([
            f"{letter}-dataset", "Baseline (Mô hình gốc)",
            float(old_m['AUC']), float(old_m['AUPR']), float(old_m['Accuracy']),
            float(old_m['Precision']), float(old_m['Recall']), float(old_m['F1-score']), float(old_m['MCC'])
        ])
        # Row for AMNTDDA
        summary_rows.append([
            f"{letter}-dataset", "AMNTDDA (Mô hình mới)",
            float(new_m['AUC']), float(new_m['AUPR']), float(new_m['Accuracy']),
            float(new_m['Precision']), float(new_m['Recall']), float(new_m['F1-score']), float(new_m['MCC'])
        ])
        
        # Calculate Improvement Row
        imp_row = [f"{letter}-dataset", "Cải thiện (%)"]
        for metric in ['AUC', 'AUPR', 'Accuracy', 'Precision', 'Recall', 'F1-score', 'MCC']:
            val_old = float(old_m[metric])
            val_new = float(new_m[metric])
            imp = ((val_new - val_old) / val_old) * 100
            imp_row.append(imp)
        summary_rows.append(imp_row)

    # Write summary rows
    current_row = 3
    for r_idx, row in enumerate(summary_rows):
        is_imp = "Cải thiện" in row[1]
        is_new = "AMNTDDA" in row[1]
        bg_color = "CCFBF1" if is_new else ("F0FDF4" if is_imp else "F8FAFC")
        text_color = "0F766E" if is_new else ("166534" if is_imp else "1E293B")
        
        ws_summary.row_dimensions[current_row].height = 22
        
        # Dataset cell (merged for every 3 rows)
        if r_idx % 3 == 0:
            ws_summary.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=1)
            cell_ds = ws_summary.cell(row=current_row, column=1, value=row[0])
            format_cell(cell_ds, bold=True, color="0F766E", fill_color="F0FDFA", alignment="center")
        
        # Format the other cells
        for col_idx, val in enumerate(row[1:], 2):
            cell = ws_summary.cell(row=current_row, column=col_idx)
            
            if isinstance(val, float):
                if is_imp:
                    cell.value = f"{val:+.2f}%"
                else:
                    cell.value = f"{val:.4f}"
            else:
                cell.value = val
                
            format_cell(cell, bold=(is_imp or is_new), color=text_color, fill_color=bg_color, alignment="center" if col_idx > 2 else "left")
            
        current_row += 1

    # Add Note section at bottom of summary
    ws_summary.merge_cells(start_row=current_row+1, start_column=1, end_row=current_row+3, end_column=9)
    note_cell = ws_summary.cell(row=current_row+1, column=1, value=(
        "GHI CHÚ THỰC NGHIỆM:\n"
        "1. Kết quả kiểm thử chéo 10 phần (10-fold cross-validation) cho cả 2 mô hình.\n"
        "2. Mô hình mới AMNTDDA tích hợp Cơ chế chú ý (Attention Mechanism) và Học tương phản đồ thị (Graph Contrastive Learning).\n"
        "3. Phần trăm cải thiện (%) dương cho thấy sự nâng cao hiệu năng rõ rệt ở hầu hết các chỉ số y sinh quan trọng (AUC, AUPR, F1-score)."
    ))
    format_cell(note_cell, font_size=9, italic=True, color="475569", fill_color="F1F5F9", alignment="left")
    ws_summary.row_dimensions[current_row+1].height = 65

    # Auto-adjust column widths
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 25

    # ────────────────────────────────────────────────────────────────────────
    # SHEETS FOR DATASETS (B, C, F)
    # ────────────────────────────────────────────────────────────────────────
    datasets_info = [
        ("B-dataset", b_old, b_new),
        ("C-dataset", c_old, c_new),
        ("F-dataset", f_old, f_new)
    ]

    for title, df_old, df_new in datasets_info:
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True
        
        # Header - Title
        ws.merge_cells("A1:T1")
        ws["A1"] = f"BẢNG KẾT QUẢ CHI TIẾT 10-FOLD TRÊN TẬP {title.upper()}"
        format_cell(ws["A1"], font_size=12, bold=True, color="FFFFFF", fill_color="0F766E", alignment="center")
        ws.row_dimensions[1].height = 30

        # Subheaders side-by-side
        # Columns A to I: Baseline Model
        # Columns K to T: AMNTDDA Model
        ws.merge_cells("A2:I2")
        ws["A2"] = "MÔ HÌNH GỐC (BASELINE)"
        format_cell(ws["A2"], font_size=10, bold=True, color="FFFFFF", fill_color="BE123C", alignment="center")
        
        ws.merge_cells("K2:T2")
        ws["K2"] = "MÔ HÌNH MỚI (AMNTDDA)"
        format_cell(ws["K2"], font_size=10, bold=True, color="FFFFFF", fill_color="0369A1", alignment="center")
        ws.row_dimensions[2].height = 22

        # Write Columns headers
        old_cols = df_old.columns.tolist()
        new_cols = df_new.columns.tolist()
        
        # Old model headers in A3:I3
        for col_idx, col_name in enumerate(old_cols, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            format_cell(cell, font_size=9, bold=True, color="334155", fill_color="FFE4E6", alignment="center")
            
        # New model headers in K3 onwards
        for col_idx, col_name in enumerate(new_cols, 11):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            format_cell(cell, font_size=9, bold=True, color="334155", fill_color="E0F2FE", alignment="center")
        ws.row_dimensions[3].height = 20

        # Write Data rows (Folds 1-10, Mean, Std)
        max_rows = max(len(df_old), len(df_new))
        
        for r_idx in range(max_rows):
            row_num = 4 + r_idx
            ws.row_dimensions[row_num].height = 18
            
            # Write Old Data in columns A-I
            if r_idx < len(df_old):
                row_data = df_old.iloc[r_idx]
                is_stats = "Mean" in str(row_data['Fold']) or "Std" in str(row_data['Fold'])
                bg_color = "FFF1F2" if is_stats else "FFFFFF"
                font_color = "9F1239" if is_stats else "1E293B"
                
                for c_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=c_idx)
                    if isinstance(val, (int, float)) and c_idx > 2:
                        cell.value = float(val)
                        cell.number_format = '0.0000'
                    else:
                        cell.value = str(val)
                    format_cell(cell, bold=is_stats, color=font_color, fill_color=bg_color, alignment="center")
            
            # Write New Data in columns K-T
            if r_idx < len(df_new):
                row_data = df_new.iloc[r_idx]
                is_stats = "Mean" in str(row_data['Fold']) or "Std" in str(row_data['Fold'])
                bg_color = "F0F9FF" if is_stats else "FFFFFF"
                font_color = "075985" if is_stats else "1E293B"
                
                for c_idx, val in enumerate(row_data, 11):
                    cell = ws.cell(row=row_num, column=c_idx)
                    if isinstance(val, (int, float)) and c_idx > 12:
                        cell.value = float(val)
                        cell.number_format = '0.0000'
                    elif c_idx == 12: # Best_Epoch is sometimes int or float
                        cell.value = float(val) if isinstance(val, (int, float)) else val
                    else:
                        cell.value = str(val)
                    format_cell(cell, bold=is_stats, color=font_color, fill_color=bg_color, alignment="center")

        # Set Column widths for sheets
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            # Empty column J spacer width
            if col_letter == 'J':
                ws.column_dimensions[col_letter].width = 4
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    # Save workbook
    wb.save(output_path)
    print(f"[OK] Excel results generated at: {output_path}")

if __name__ == "__main__":
    main()
